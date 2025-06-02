import pandas as pd
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from config.constants import COLUMN_TYPE_MAPPING, COLUMN_TYPE_OUTPUT_CN, COLUMN_TYPE_OUTPUT_JP

def data_sort_func(selected_df, internal_column_names, prices, output_sheets_config, save_path):
    # 1. 列重命名
    df = selected_df.copy()
    df.columns = internal_column_names
    # 2. 删除全为空的行
    df.dropna(how='all', inplace=True)
    # 3. 验证 count_* 列
    count_cols = [col for col in ["count_ani", "count_coloring", "count_1_yuan", "count_2_yuan"] if col in df.columns]
    error_rows = []
    for col in count_cols:
        original_col = df[col]
        numeric_col = pd.to_numeric(original_col, errors='coerce')
        for idx in df.index:
            original_value = original_col[idx]
            numeric_value = numeric_col[idx]
            if pd.isna(numeric_value):
                if pd.isna(original_value) or str(original_value).strip() == '':
                    reason = "为空值"
                else:
                    reason = f"无法解析为数字（原值：{original_value!r}）"
                error_rows.append(f"第 {idx + 2} 行（列名：{col}）的值非法：{reason}")
            elif numeric_value < 0:
                error_rows.append(f"第 {idx + 2} 行（列名：{col}）的值非法：是负数（原值：{original_value}）")
    if error_rows:
        raise ValueError("检测到以下单元格不是合法的非负数字：\n" + "\n".join(error_rows))
    # 校验通过后再安全转换
    for col in count_cols:
        df[col] = pd.to_numeric(df[col], errors='raise').astype(float)
    # 4. 排序
    sort_keys = [k for k in ["animation_name", "animation_episode", "order_number"] if k in df.columns]
    df.sort_values(by=sort_keys, ascending=True, inplace=True)
    # 5. 写入 Excel（检查点工作表）
    df_sorted_for_excel = df.rename(columns=COLUMN_TYPE_OUTPUT_CN)

    # 插入 data_sorted_sum 工作表（带合计行）
    sum_group_keys_all = ["株式会社", "片名", "话数"]
    actual_sum_group_keys = [col for col in sum_group_keys_all if col in df_sorted_for_excel.columns]
    sum_value_fields = ["动画", "上色", "一原", "二原"]
    actual_value_fields = [col for col in sum_value_fields if col in df_sorted_for_excel.columns]
    if actual_sum_group_keys:
        grouped = df_sorted_for_excel.groupby(actual_sum_group_keys, as_index=False)
        rows_with_sums = []
        for _, group in grouped:
            rows_with_sums.extend(group.to_dict('records'))
            sum_row = {col: "" for col in df_sorted_for_excel.columns}
            sum_row["传票号"] = "SUM"
            for key in actual_sum_group_keys:
                sum_row[key] = group.iloc[0][key]
            for field in actual_value_fields:
                sum_row[field] = group[field].sum()
            rows_with_sums.append(sum_row)
        df_sorted_with_sum = pd.DataFrame(rows_with_sums, columns=df_sorted_for_excel.columns)

    # 1. 删除 order_number 列
    df_for_calc = df.drop(columns=["order_number"]) if "order_number" in df.columns else df.copy()
    # 2. 动态选取存在的 count_* 列
    count_cols = [col for col in ["count_ani", "count_coloring", "count_1_yuan", "count_2_yuan"] if
                  col in df_for_calc.columns]
    # 3. 分组汇总
    # 预设想要用来分组的列
    desired_group_keys = ["company_name", "animation_name", "animation_episode"]
    # 实际存在于 DataFrame 中的分组列
    actual_group_keys = [col for col in desired_group_keys if col in df_for_calc.columns]
    # 如果没有分组列，直接使用原始数据；否则进行分组汇总
    if actual_group_keys:
        df_for_calc = df_for_calc.groupby(actual_group_keys, as_index=False)[count_cols].sum()
    # 4. 转换回中文列名
    df_calculated_for_excel = df_for_calc.rename(columns=COLUMN_TYPE_OUTPUT_CN)
    # 5. 添加价格列
    price_field_map = {
        cn_name: internal_name
        for cn_name, internal_name in COLUMN_TYPE_MAPPING.items()
        if internal_name and internal_name.startswith("price_")
    }
    for cn_label, internal_col in price_field_map.items():
        if cn_label in prices:
            df_for_calc[internal_col] = prices[cn_label]
    # 6. 计算总价列
    for count_col, price_col, total_col in [
        ("count_ani", "price_ani", "total_ani"),
        ("count_coloring", "price_coloring", "total_coloring"),
        ("count_1_yuan", "price_1_yuan", "total_1_yuan"),
        ("count_2_yuan", "price_2_yuan", "total_2_yuan"),
    ]:
        if count_col in df_for_calc.columns and price_col in df_for_calc.columns:
            df_for_calc[total_col] = df_for_calc[count_col] * df_for_calc[price_col]


    # 导出Excel
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        align_center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        # 写入动态输出工作表
        used_sheet_names = set()
        default_base_name = "工作表"
        name_counter = 1
        for sheet in output_sheets_config:
            if not sheet.get("enabled"):
                continue
            style = sheet.get("style", "中文")
            sheet_name = sheet.get("sheet_name", "").strip()
            # 如果用户没有填写 sheet_name，就自动生成唯一的默认名
            if not sheet_name:
                while True:
                    generated_name = f"{default_base_name}{name_counter}"
                    name_counter += 1
                    if generated_name not in used_sheet_names:
                        sheet_name = generated_name
                        break
            # 如果填写了，确保与之前的不冲突
            elif sheet_name in used_sheet_names:
                suffix = 1
                original_name = sheet_name
                while sheet_name in used_sheet_names:
                    sheet_name = f"{original_name}_{suffix}"
                    suffix += 1
            # 更新工作表名并记录已使用的名称
            sheet["sheet_name"] = sheet_name
            used_sheet_names.add(sheet_name)
            # 提取所有请求的列
            columns = [col for col in sheet.get("columns", []) if col]
            # 新建一个 DataFrame，逐列添加（无则填 0）
            df_to_output = pd.DataFrame()
            for col in columns:
                if col in df_for_calc.columns:
                    df_to_output[col] = df_for_calc[col]
                else:
                    df_to_output[col] = 0  # 不存在的列统一填 0
            # 识别出当前表中的数量列（英文字段名）
            count_fields_in_this_sheet = [
                col for col in ["count_ani", "count_coloring", "count_1_yuan", "count_2_yuan"]
                if col in df_to_output.columns
            ]
            # 若存在这些列，且某行这些列全为0，则删除该行
            if count_fields_in_this_sheet:
                non_zero_mask = (df_to_output[count_fields_in_this_sheet] != 0).any(axis=1)
                df_to_output = df_to_output[non_zero_mask]
            # 应用列名样式
            if style == "中文":
                df_to_output.rename(columns=COLUMN_TYPE_OUTPUT_CN, inplace=True)
            elif style == "日文":
                df_to_output.rename(columns=COLUMN_TYPE_OUTPUT_JP, inplace=True)
            # 写入工作表
            df_to_output.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            # 设置样式
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                           min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    col_letter = get_column_letter(cell.column)
                    col_name = worksheet[f"{col_letter}1"].value
                    if col_name in ["株式会社", "片名", "タィトル"]:
                        worksheet.column_dimensions[col_letter].width = 50
                    if col_name in ["话数", "动画", "上色", "一原", "二原", "動画", "仕上げ", "L/Oー作监", "2原", "単価", "合計"]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
        # 写入 calculated 表
        df_calculated_for_excel.to_excel(writer, sheet_name='data_calculated', index=False)
        worksheet_calculated = writer.sheets['data_calculated']
        # 设置格式
        for row in worksheet_calculated.iter_rows(min_row=1, max_row=worksheet_calculated.max_row,
                                       min_col=1, max_col=worksheet_calculated.max_column):
            for cell in row:
                col_letter = get_column_letter(cell.column)
                col_name = worksheet_calculated[f"{col_letter}1"].value
                # 设置列宽
                if worksheet_calculated[f"{col_letter}1"].row == 1:
                    if col_name in ["株式会社", "片名"]:
                        worksheet_calculated.column_dimensions[col_letter].width = 50
                # 居中对齐特定列
                if col_name in ["话数", "动画", "上色", "一原", "二原"]:
                    cell.alignment = align_center
                # 所有格子添加边框
                cell.border = thin_border
        # 写入 sorted_sum 表
        df_sorted_with_sum.to_excel(writer, sheet_name="data_sorted_sum", index=False)
        worksheet_sorted_sum = writer.sheets["data_sorted_sum"]
        bold_font = Font(bold=True)
        # 设置格式
        for row in worksheet_sorted_sum.iter_rows(min_row=1, max_row=worksheet_sorted_sum.max_row,
                                       min_col=1, max_col=worksheet_sorted_sum.max_column):
            for cell in row:
                col_letter = get_column_letter(cell.column)
                col_name = worksheet_sorted_sum[f"{col_letter}1"].value
                # 设置列宽
                if worksheet_sorted_sum[f"{col_letter}1"].row == 1:
                    if col_name in ["株式会社", "片名"]:
                        worksheet_sorted_sum.column_dimensions[col_letter].width = 50
                # 居中对齐特定列
                if col_name in ["话数", "动画", "上色", "一原", "二原"]:
                    cell.alignment = align_center
                # 所有格子添加边框
                cell.border = thin_border
        # 汇总行加粗
        for row in worksheet_sorted_sum.iter_rows(min_row=2, max_row=worksheet_sorted_sum.max_row):
            first_cell_value = row[0].value
            if first_cell_value == "SUM":
                for cell in row:
                    cell.font = bold_font
        # 写入 sorted 表
        df_sorted_for_excel.to_excel(writer, sheet_name='data_sorted', index=False)
        worksheet_sorted = writer.sheets['data_sorted']
        # 设置格式
        for row in worksheet_sorted.iter_rows(min_row=1, max_row=worksheet_sorted.max_row,
                                       min_col=1, max_col=worksheet_sorted.max_column):
            for cell in row:
                col_letter = get_column_letter(cell.column)
                col_name = worksheet_sorted[f"{col_letter}1"].value
                # 设置列宽
                if worksheet_sorted[f"{col_letter}1"].row == 1:
                    if col_name in ["株式会社", "片名"]:
                        worksheet_sorted.column_dimensions[col_letter].width = 50
                # 居中对齐特定列
                if col_name in ["话数", "动画", "上色", "一原", "二原"]:
                    cell.alignment = align_center
                # 所有格子添加边框
                cell.border = thin_border

    return